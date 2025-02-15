"""
Script: SkyPrep Data Migration Tool
Description: This tool provides functionalities for transferring data
             from ADP reports to SkyPrep Bulk update templates
             using an interactive GUI interface.
Version: 1.0
Date: 2024-12-20
Developer: Saikat Datta
"""
# region Imports
# -----------------------------------------------------------
# Imports Section
# Handles all library and module imports required for the script
# -----------------------------------------------------------
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
import openpyxl.styles
import pandas as pd
from datetime import datetime
import logging
#endregion

# region Clean Report
# -----------------------------------------------------------
# Clean Report Section
# Handles the cleansing of data, applying rules, and saving
# the cleaned report to a new Excel file.
# -----------------------------------------------------------
clean_file_path = ""

# Browse and select an Excel file
def select_clean_file():
    global clean_file_path
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        clean_file_path = file_path
        file_label.config(text=f"Selected File: {os.path.basename(file_path)}")
    else:
        file_label.config(text="No file selected")

# Read the uploaded Excel file, process it, and save the result
def start_clean_logic():
    global clean_file_path
    if not clean_file_path:
        messagebox.showerror("Error", "Please upload an Excel file before starting.")
        return

    try:
        # Determine the selected report
        report_type = selected_report.get()
        messagebox.showinfo("Selected Report", f"Processing: {report_type} Report")

        # Create the progress bar
        progress_bar = ttk.Progressbar(bottom_bar, orient="horizontal", mode="determinate", length=400)
        progress_bar.pack(pady=5)
        
        if report_type == "All_Course_Progresses":
            # Handle Duplicate Removal logic
            data_frame = pd.read_excel(clean_file_path)
            data_frame["Email_Course"] = data_frame["Email"] + " | " + data_frame["Course Name"]
            data_frame = data_frame.sort_values(by=["Email_Course", "Start Date", "Completion Date", "Expiration Date"], ascending=[True, False, False, False])
            data_frame_cleaned = data_frame.drop_duplicates(subset=["Email_Course"], keep="first")
            data_frame_cleaned = data_frame_cleaned.drop(columns=["Email_Course"])
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Cleaned Data",
                initialfile="Output_ADP_All_Course_Progresses_Report_Cleaned.xlsx"
            )
            if not save_path:
                messagebox.showinfo("Cancelled", "Save operation was cancelled.")
                progress_bar.pack_forget()
                return
            data_frame_cleaned.to_excel(save_path, index=False)
            messagebox.showinfo("Success", f"Cleaned data saved to: {save_path}")

        elif report_type == "Deficiency_Recertification":
            # Handle Deficiency Recertification logic
            wb = openpyxl.load_workbook(clean_file_path)
            sheet = wb.active
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active

            required_columns = [
                "Position ID", "Payroll Name", "Course Name Description",
                "Start Date", "Recertification Date", "Acquired Date",
            ]
            headers = [cell.value for cell in sheet[1]]
            required_indices = [headers.index(col) for col in required_columns]

            new_sheet.append(required_columns)
            total_rows = sheet.max_row - 1
            progress_bar["maximum"] = total_rows

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                progress_bar["value"] = idx
                progress_bar.update()
                row_list = list(row)
                filtered_row = [row_list[idx] for idx in required_indices]

                start_date = filtered_row[required_columns.index("Start Date")]
                recertification_date = filtered_row[required_columns.index("Recertification Date")]
                acquired_date = filtered_row[required_columns.index("Acquired Date")]

                if start_date and not recertification_date and not acquired_date:
                    pass
                elif start_date and acquired_date and not recertification_date:
                    filtered_row[required_columns.index("Recertification Date")] = None
                    filtered_row[required_columns.index("Acquired Date")] = None
                elif start_date and recertification_date:
                    if recertification_date > start_date:
                        filtered_row[required_columns.index("Acquired Date")] = start_date
                    elif recertification_date == start_date:
                        filtered_row[required_columns.index("Recertification Date")] = None
                        filtered_row[required_columns.index("Acquired Date")] = None
                    elif recertification_date < start_date:
                        filtered_row[required_columns.index("Recertification Date")] = None
                        filtered_row[required_columns.index("Acquired Date")] = None

                new_sheet.append(filtered_row)

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Cleaned Data",
                initialfile="Output_ADP_Deficiency_Recertification_Report_Cleaned.xlsx"
            )
            if not save_path:
                messagebox.showinfo("Cancelled", "Save operation was cancelled.")
                progress_bar.pack_forget()
                return
            new_wb.save(save_path)
            messagebox.showinfo("Success", f"Cleaned data saved to: {save_path}")

        elif report_type == "Policies_Certifications_Vaccines_Licences":
            # Handle Policies, Certifications, Vaccines and Licenses logic
            wb = openpyxl.load_workbook(clean_file_path)
            sheet = wb.active
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active

            existing_columns = [
                "Position ID", "Payroll Name", "License/Certification Description",
                "Effective Date", "Expiration Date", "Hire Date",
            ]
            required_columns = [
                "Position ID", "Payroll Name", "Course Name Description",
                "Start Date", "Recertification Date", "Acquired Date",
            ]
            new_sheet.append(required_columns)

            existing_headers = [cell.value for cell in sheet[1]]
            existing_indices = [existing_headers.index(col) for col in existing_columns]

            total_rows = sheet.max_row - 1
            progress_bar["maximum"] = total_rows

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                progress_bar["value"] = idx
                progress_bar.update()
                row_list = list(row)
                filtered_row = [row_list[idx] for idx in existing_indices]

                position_id = filtered_row[existing_columns.index("Position ID")]
                payroll_name = filtered_row[existing_columns.index("Payroll Name")]
                course_name_description = filtered_row[existing_columns.index("License/Certification Description")]
                
                start_date = filtered_row[existing_columns.index("Effective Date")]
                recertification_date = filtered_row[existing_columns.index("Expiration Date")]
                hire_date = filtered_row[existing_columns.index("Hire Date")]

                if start_date == None:
                    if recertification_date == None:
                        start_date = hire_date
                        acquired_date = None
                    else:
                        start_date = hire_date
                        acquired_date = start_date
                else:
                    acquired_date = start_date
                    if recertification_date == None:
                        recertification_date = datetime(2050, 1, 1)

                if recertification_date == hire_date:
                    acquired_date == None
                    recertification_date == None

                # Prepare the row for the new sheet
                transformed_row = [
                    position_id or "", payroll_name or "", course_name_description or "",
                    start_date or "", recertification_date or "", acquired_date or ""
                ]
                new_sheet.append(transformed_row)

            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Cleaned Data",
                initialfile="Output_ADP_Policies_Certifications_Vaccines_Licences_Report_Cleaned.xlsx"
            )
            if not save_path:
                messagebox.showinfo("Cancelled", "Save operation was cancelled.")
                progress_bar.pack_forget()
                return
            new_wb.save(save_path)
            messagebox.showinfo("Success", f"Cleaned data saved to: {save_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        # Remove the progress bar after completion
        progress_bar.pack_forget()
# endregion

# region Transform Report
# -----------------------------------------------------------
# Transform Report Section
# Manages the transformation of data, including mapping fields,
# modifying structures and preparing the transformed report.
# -----------------------------------------------------------
transform_file_path = ""
course_mapping_file_path = ""
user_list_file_path = ""

def select_transform_file():
    """Browse and select the main Excel file for transformation."""
    global transform_file_path
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        transform_file_path = file_path
        transform_file_label.config(text=f"Selected File: {os.path.basename(file_path)}")
    else:
        transform_file_label.config(text="No file selected")

def select_course_mapping_file():
    """Browse and select the course mapping Excel file."""
    global course_mapping_file_path
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        course_mapping_file_path = file_path
        course_mapping_file_label.config(text=f"Selected File: {os.path.basename(file_path)}")
    else:
        course_mapping_file_label.config(text="No file selected")

def select_user_list_file():
    """Browse and select the user list Excel file."""
    global user_list_file_path
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        user_list_file_path = file_path
        user_list_file_label.config(text=f"Selected File: {os.path.basename(file_path)}")
    else:
        user_list_file_label.config(text="No file selected")

def start_transform_logic():
    """Perform the transformation logic as per the requirements."""
    if not (transform_file_path and course_mapping_file_path and user_list_file_path):
        messagebox.showerror("Error", "Please upload all required files.")
        return

    try:
        # Create the progress bar
        progress_bar = ttk.Progressbar(bottom_bar, orient="horizontal", mode="determinate", length=400)
        progress_bar.pack(pady=5)

        # Open the main Excel file
        main_wb = openpyxl.load_workbook(transform_file_path)
        main_sheet = main_wb.active

        # Open the course mapping Excel file
        course_mapping_wb = openpyxl.load_workbook(course_mapping_file_path)
        course_mapping_sheet = course_mapping_wb.active

        # Open the user list Excel file
        user_list_wb = openpyxl.load_workbook(user_list_file_path)
        user_list_sheet = user_list_wb.active

        # Create a new workbook for the transformed data
        transformed_wb = openpyxl.Workbook()
        transformed_sheet = transformed_wb.active
        transformed_sheet.title = "Transformed Data"

        # Create a separate sheet for "Discarded Data"
        discarded_sheet = transformed_wb.create_sheet(title="Discarded Data")

        # Create a separate sheet for rows with "Login Status: Not found"
        not_found_sheet = transformed_wb.create_sheet(title="Not Found Records")

        # Define the mapping for headers between the main file and transformed sheet
        main_to_transformed_mapping = {
            "Position ID": "Work phone",
            "Course Name Description": "Course Name",
            "Start Date": "Start Date",
            "Recertification Date": "Expiration Date",
            "Acquired Date": "Completion Date",
        }

        # Define additional static fields for the transformed sheet
        additional_fields = {
            "Login Status": lambda email: "Active" if email else "Not found",
            "Course Progress Status": lambda recertification_date: "passed" if recertification_date else "not-started",
            "Deadline Date": lambda: "",  # Always blank
        }

        # Write the headers to the transformed sheet
        transformed_headers = [
            "SkyPrep ID", "First name", "Last name", "Email",
            "Work phone", "Course Number", "Course Name",
            "Login Status", "Course Progress Status",
            "Start Date", "Completion Date",
            "Deadline Date", "Expiration Date"
        ]
        transformed_sheet.append(transformed_headers)

        # Extract headers from the main file
        main_headers = [cell.value for cell in main_sheet[1]]

        # Map main headers to their indices
        main_header_indices = {header: idx for idx, header in enumerate(main_headers)}

        # Ensure all required headers are present in the main file
        missing_headers = [
            header for header in main_to_transformed_mapping.keys()
            if header not in main_header_indices
        ]
        if missing_headers:
            messagebox.showerror("Error", f"Missing required columns in main file: {', '.join(missing_headers)}")
            return

        # Write the headers to the Not Found Records sheet
        no_records_headers = ["Position ID", "Payroll Name", "Login Status"]
        not_found_sheet.append(no_records_headers)

        # Write headers from the original source file to the "Discarded Data" sheet
        discarded_headers = main_headers  # Same headers as the headers from main file
        discarded_sheet.append(discarded_headers)

        # Create a set to track unique position IDs in the Not Found Records sheet
        existing_position_ids = set()
        
        # Initialize progress bar
        total_rows = main_sheet.max_row - 1  # Exclude the header row
        progress_bar["maximum"] = total_rows
        
        # Extract headers from the user list file
        user_list_headers = [cell.value for cell in user_list_sheet[1]]

        # Map user list headers to their indices
        user_list_header_indices = {header: idx for idx, header in enumerate(user_list_headers)}

        # Process rows in the main file
        for idx, row in enumerate(main_sheet.iter_rows(min_row=2, values_only=True), start=1):
            # Update progress bar
            progress_bar["value"] = idx
            progress_bar.update()

            # Extract data from the main sheet
            position_id = row[main_header_indices.get("Position ID")]
            payroll_name = row[main_header_indices.get("Payroll Name")]
            course_name_description = row[main_header_indices.get("Course Name Description")]
            start_date = row[main_header_indices.get("Start Date")]
            recertification_date = row[main_header_indices.get("Recertification Date")]
            acquired_date = row[main_header_indices.get("Acquired Date")]

            # Perform course mapping
            course_number_skyprep = None
            course_name_skyprep = None
            for mapping_row in course_mapping_sheet.iter_rows(min_row=2, values_only=True):
                if mapping_row[0] == course_name_description:
                    course_number_skyprep = mapping_row[1]
                    course_name_skyprep = mapping_row[2]
                    break

            # If course is marked as "Discard", store it in the Discarded Data sheet
            if course_name_skyprep == "Discard":
                discarded_sheet.append(list(row))
                continue
            
            # Check if course mapping not found
            elif course_name_skyprep == None:
                course_name_skyprep = "Course Mapping Not Found"
            
            # Perform user mapping
            skyprep = email = first_name = last_name = None
            for user_row in user_list_sheet.iter_rows(min_row=2, values_only=True):
                if user_row[user_list_header_indices["work_phone"]] == position_id:
                    skyprep = user_row[user_list_header_indices["skyprep_internal_id"]]
                    email = user_row[user_list_header_indices["email_or_username"]]
                    first_name = user_row[user_list_header_indices["first_name"]]
                    last_name = user_row[user_list_header_indices["last_name"]]
                    break

            # Determine additional fields
            login_status = additional_fields["Login Status"](email)
            course_progress_status = additional_fields["Course Progress Status"](recertification_date)
            deadline_date = additional_fields["Deadline Date"]()

            # Remove start date if course progress status is not started
            if course_progress_status == "not-started":
                start_date = None

            # Append to the appropriate sheet
            if login_status == "Not found":
                # Prepare the row for the records not found sheet
                no_records_row = [position_id or "", payroll_name or "", login_status]
                # Check if the position_id already exists in the set
                if position_id not in existing_position_ids:
                    not_found_sheet.append(no_records_row)
                    existing_position_ids.add(position_id)  # Add to the set after appending
            else:
                # Prepare the row for the transformed sheet
                transformed_row = [
                    skyprep or "", first_name or "", last_name or "",
                    email or "", position_id or "",
                    course_number_skyprep or "", course_name_skyprep or "",
                    login_status, course_progress_status,
                    start_date or "", acquired_date or "",
                    deadline_date, recertification_date or ""
                ]
                transformed_sheet.append(transformed_row)

        # Ask the user where to save the transformed file
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Transformed Data",
            initialfile="Output_ADP_All_Course_Progresses_Report.xlsx"
        )
        if not save_path:
            messagebox.showinfo("Cancelled", "Save operation was cancelled.")
            progress_bar.pack_forget()  # Remove progress bar on cancel
            return

        # Save the transformed workbook
        transformed_wb.save(save_path)
        messagebox.showinfo("Success", f"Transformed data saved to: {save_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        # Remove the progress bar after completion
        progress_bar.pack_forget()
# endregion

# region Transfer Report
# -----------------------------------------------------------
# Transfer Report Section
# Facilitates data transfer operations, including organizing
# data into the target format and saving the result.
# -----------------------------------------------------------
transfer_file_path = ""

def select_transfer_file():
    """Select an Excel file for the Transfer section."""
    global transfer_file_path
    file_path = filedialog.askopenfilename(
        title="Select an Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
    )
    if file_path:
        transfer_file_path = file_path
        transfer_file_label.config(text=f"Selected File: {os.path.basename(file_path)}")
    else:
        transfer_file_label.config(text="No file selected")

def generate_destination_columns(max_courses=84):
    """Dynamically generate destination columns."""
    columns = ['skyprep_internal_id', 'first_name', 'last_name', 'email_or_username', 'work_phone']
    for i in range(1, max_courses + 1):
        columns.extend([
            f'course {i}', f'course {i} status', f'course {i} date started', f'course {i} date finished',
            f'course {i} access date', f'course {i} deadline date', f'course {i} expiration date'
        ])
    return columns

def start_transfer_logic():
    """Transfer the source data into the desired format."""
    if not (transfer_file_path):
        messagebox.showerror("Error", "Please upload an Excel file before starting.")
        return
    try:
        # Create the progress bar
        progress_bar = ttk.Progressbar(bottom_bar, orient="horizontal", mode="determinate", length=400)
        progress_bar.pack(pady=5)

        # Load the source file
        source_data_frame = pd.read_excel(transfer_file_path)

        # Generate destination columns dynamically
        destination_columns = generate_destination_columns()

        # Initialize a list to collect rows
        rows_list = []

        # Create an empty DataFrame with the destination format columns
        output_data_frame = pd.DataFrame(columns=destination_columns)
        grouped = source_data_frame.groupby('SkyPrep ID')

        # Set progress bar maximum to the number of groups
        total_groups = len(grouped)
        progress_bar["maximum"] = total_groups

        for idx, (employee, group) in enumerate(grouped, start=1):
            row = {col: '' for col in destination_columns}
            row['skyprep_internal_id'] = employee
            row['first_name'] = group['First name'].iloc[0]
            row['last_name'] = group['Last name'].iloc[0]
            row['email_or_username'] = group['Email'].iloc[0]
            row['work_phone'] = group['Work phone'].iloc[0]

            for _, course in group.iterrows():
                course_number = course['Course Number']
                course_name = course['Course Name']
                course_progress_status = course['Course Progress Status']
                start_date = course['Start Date']
                completion_date = course['Completion Date']
                expiration_date = course['Expiration Date']

                for i in range(1, (len(destination_columns) - 5) // 7 + 1): #Static Columns=5, Dynamic Columns=7
                    target_course_column = f'course {i}'
                    if target_course_column in destination_columns and course_number == f'Course {i}':
                        row[target_course_column] = course_name
                        row[f'course {i} status'] = course_progress_status
                        row[f'course {i} date started'] = start_date
                        row[f'course {i} date finished'] = completion_date
                        row[f'course {i} expiration date'] = expiration_date
                        break

            # Add the row to the list
            rows_list.append(row)

            # Update the progress bar
            progress_bar["value"] = idx
            progress_bar.update()
            
        # After processing all rows, create the final DataFrame
        output_data_frame = pd.DataFrame(rows_list, columns=destination_columns)

        # Save the transformed data to a new file
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Transformed File",
            initialfile="Output_ADP_Bulk_Update_User_List (including courses).xlsx"
        )
        if output_file_path:
            output_data_frame.to_excel(output_file_path, index=False, engine='openpyxl')
            messagebox.showinfo("Success", f"File saved successfully:\n{output_file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        # Remove the progress bar after completion
        progress_bar.pack_forget()
# endregion

# region Compare Report
# -----------------------------------------------------------
# Compare Report Section
# Facilitates data comparison between the generated user list
# and the bulk update user list downloaded from SkyPrep.
# -----------------------------------------------------------
compare_file_path = ""
reference_file_path = ""

def select_compare_file():
    """Select the Compare Excel file."""
    global compare_file_path
    file_path = filedialog.askopenfilename(
        title="Select Compare Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
    )
    if file_path:
        compare_file_path = file_path
        compare_file_label.config(text=f"Compare File: {os.path.basename(file_path)}")
    else:
        compare_file_label.config(text="No file selected")

def select_reference_file():
    """Select the Reference Excel file."""
    global reference_file_path
    file_path = filedialog.askopenfilename(
        title="Select Reference Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
    )
    if file_path:
        reference_file_path = file_path
        reference_file_label.config(text=f"Reference File: {os.path.basename(file_path)}")
    else:
        reference_file_label.config(text="No file selected")

def start_compare_logic():
    """Compare the uploaded sheets and update values based on the comparison."""
    if not (compare_file_path and reference_file_path):
        messagebox.showerror("Error", "Please upload both files for comparison.")
        return
    
    # Define the log file name
    log_file = "update_log.txt"

    # Write the header before setting up logging
    with open(log_file, "w") as log:
        log.write(
            "Skyprep_ID,"
            "Last_Name,"
            "First_Name,"
            "Course_ID,"
            "Course_Name(SkyPrep),"
            "Final_Status,"
            "Final_Start_Date,"
            "Final_Finish_Date,"
            "Final_Expiration_Date,"
            "SkyPrep_Status,"
            "SkyPrep_Start_Date,"
            "SkyPrep_Finish_Date,"
            "SkyPrep_Expiration_Date,"
            "ADP_Status,"
            "ADP_Start_Date,"
            "ADP_Finish_Date,"
            "ADP_Expiration_Date,"
            "Row_Number,"
            "Timestamp\n")
    
    # Configure logging to write to a file
    logging.basicConfig(
        filename=log_file,
        filemode="a", # Append mode to retain the header
        format="%(message)s,%(asctime)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        level=logging.INFO
    )

    try:
        # Create a progress bar
        progress_bar = ttk.Progressbar(bottom_bar, orient="horizontal", mode="determinate", length=400)
        progress_bar.pack(pady=5)

        # Load the Compare and Reference workbooks
        compare_wb = openpyxl.load_workbook(compare_file_path)
        reference_wb = openpyxl.load_workbook(reference_file_path)
        
        # Assume the first sheet is the active one in both files
        compare_sheet = compare_wb.active
        reference_sheet = reference_wb.active

        # Get the headers from both sheets
        compare_headers = [cell.value for cell in compare_sheet[1]]
        reference_headers = [cell.value for cell in reference_sheet[1]]

        # Define the key column for matching rows and declare the total number of courses
        key_column = "skyprep_internal_id"
        max_courses = 84

        # Find the index of the key column in both sheets
        compare_key_idx = compare_headers.index(key_column)
        reference_key_idx = reference_headers.index(key_column)

        # Initialize progress bar
        total_rows = compare_sheet.max_row - 1  # Exclude the header row
        progress_bar["maximum"] = total_rows

        # Loop through each row in the Compare sheet (starting from the second row)
        for compare_row_idx, compare_row in enumerate(compare_sheet.iter_rows(min_row=2, values_only=True), start=2):
            compare_key = compare_row[compare_key_idx]
            compare_last_name = compare_row[2]
            compare_first_name = compare_row[1]

            # Search for the matching key in the Reference sheet
            for reference_row in reference_sheet.iter_rows(min_row=2, values_only=True):
                if reference_row[reference_key_idx] == compare_key:
                    
                    # Match found - loop through all the courses
                    for i in range(1, (max_courses + 1)):
                        # Define course column group names dynamically
                        column_names = [
                            f"course {i}",
                            f"course {i} status",
                            f"course {i} date started",
                            f"course {i} date finished",
                            f"course {i} deadline date",
                            f"course {i} expiration date",
                        ]

                        # Check if these columns exist in both sheets
                        if all(col in compare_headers and col in reference_headers for col in column_names):
                            # Get column indices dynamically
                            compare_indices = {name: compare_headers.index(name) for name in column_names}
                            reference_indices = {name: reference_headers.index(name) for name in column_names}

                            # Extract values from Compare and Reference rows
                            compare_values = {name: compare_row[idx] for name, idx in compare_indices.items()}
                            reference_values = {name: reference_row[idx] for name, idx in reference_indices.items()}

                            # Get course status
                            compare_course_status = compare_values[f"course {i} status"]
                            reference_course_status = reference_values[f"course {i} status"]

                            # Get course dates
                            compare_date_started = compare_values[f"course {i} date started"]
                            compare_date_finished = compare_values[f"course {i} date finished"]
                            compare_expiration_date = compare_values[f"course {i} expiration date"]

                            reference_date_started = reference_values[f"course {i} date started"]
                            reference_date_finished = reference_values[f"course {i} date finished"]
                            reference_deadline_date = reference_values[f"course {i} deadline date"]
                            reference_expiration_date = reference_values[f"course {i} expiration date"]

                            # Variables for logging purpose only
                            adp_course_status = compare_course_status
                            adp_date_started = compare_date_started
                            adp_date_finished = compare_date_finished
                            adp_expiration_date = compare_expiration_date

                            skyprep_course_status = reference_course_status
                            skyprep_date_started = reference_date_started
                            skyprep_date_finished = reference_date_finished
                            skyprep_expiration_date = reference_expiration_date

                            # Skip this course if course {i} in the Compare file is None
                            if compare_values[f"course {i}"] is not None:

                                # Initialize update needed as false
                                update_needed = False

                                # Condition 1: If course status is 'passed' in the compare sheet
                                if compare_course_status == "passed":
                                    if reference_course_status == "passed":
                                        if (reference_date_started is None) and (reference_date_finished is not None):
                                            reference_date_started = reference_date_finished
                                        elif (reference_date_started is not None) and (reference_date_finished is None):
                                            reference_date_finished = reference_date_started
                                        elif (reference_date_started is None) and (reference_date_finished is None):
                                            reference_date_started = compare_date_started
                                            reference_date_finished = compare_date_finished
                                            reference_expiration_date = compare_expiration_date

                                        if reference_expiration_date is None:
                                            if compare_expiration_date.strftime("%Y") == "2050":
                                                reference_expiration_date = compare_expiration_date
                                            else:
                                                reference_expiration_date = reference_date_finished + (compare_expiration_date - compare_date_finished)

                                        if reference_date_started.strftime("%Y-%m-%d") == compare_date_started.strftime("%Y-%d-%m"):
                                            update_needed = False
                                        elif reference_date_finished > compare_date_finished:
                                            compare_values[f"course {i} date started"] = reference_date_started
                                            compare_values[f"course {i} date finished"] = reference_date_finished
                                            compare_values[f"course {i} expiration date"] = reference_expiration_date
                                        
                                            update_needed = True
                                    else:
                                        update_needed = False

                                # Condition 2: If course status is 'not-started' in the compare sheet
                                elif compare_course_status == "not-started":
                                    if (reference_course_status == "passed"):
                                        if reference_date_started is None and reference_date_finished is not None:
                                            reference_date_started = reference_date_finished
                                        elif reference_date_started is not None and reference_date_finished is None:
                                            reference_date_finished = reference_date_started
                                        
                                        compare_values[f"course {i} status"] = reference_course_status
                                        compare_values[f"course {i} date started"] = reference_date_started
                                        compare_values[f"course {i} date finished"] = reference_date_finished
                                        compare_values[f"course {i} expiration date"] = reference_expiration_date

                                        update_needed = True
                                    
                                    elif (reference_course_status == "in-progress"):
                                        compare_values[f"course {i} status"] = reference_course_status
                                        compare_values[f"course {i} date started"] = reference_date_started
                                        compare_values[f"course {i} deadline date"] = reference_deadline_date

                                        update_needed = True

                                    else:
                                        update_needed = False                                        

                                if update_needed == True:
                                    # Update Compare Sheet
                                    for key in ["status", "date started", "date finished", "deadline date", "expiration date"]:
                                        col_name = f"course {i} {key}"
                                        compare_sheet.cell(row=compare_row_idx, column=compare_indices[col_name] + 1).value = compare_values[col_name]

                                # Log the update
                                logging.info(
                                    f"{compare_key},{compare_last_name},{compare_first_name},"
                                    f"Course {i},{compare_values[f'course {i}']},"
                                    f"{compare_values[f'course {i} status']},"
                                    f"{compare_values[f'course {i} date started']},"
                                    f"{compare_values[f'course {i} date finished']},"
                                    f"{compare_values[f'course {i} expiration date']},"
                                    f"{skyprep_course_status},{skyprep_date_started},"
                                    f"{skyprep_date_finished},{skyprep_expiration_date},"
                                    f"{adp_course_status},{adp_date_started},"
                                    f"{adp_date_finished},{adp_expiration_date},"
                                    f"{compare_row_idx}"
                                )
            
            # Update the progress bar
            progress_bar["value"] = compare_row_idx - 1  # Adjust for 1-based indexing
            progress_bar.update()
        
        # Save the updated Compare workbook
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Updated Compare File",
            initialfile="Final_Bulk_Update_File.xlsx"
        )
        if output_file_path:
            compare_wb.save(output_file_path)
            messagebox.showinfo("Success", f"Updated Compare File saved to: {output_file_path}")
        else:
            messagebox.showinfo("Cancelled", "Save operation was cancelled.")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        # Remove the progress bar after completion
        progress_bar.pack_forget()
# endregion

# region Main Window
# -----------------------------------------------------------
# Main Window Section
# Sets up the main GUI window, including navigation buttons,
# layout configuration and frame management.
# -----------------------------------------------------------

root = tk.Tk()
root.title("SkyPrep Migration Tool")
root.geometry("600x400")
root.minsize(600, 400)  # Set minimum size to prevent distortion
root.configure(bg="#2E2E2E")  # Background color

# Set favicon path for logo at the top left corner of the frame
icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "favicon.ico")
root.iconbitmap(icon_path)

# Main container frames
main_frame = tk.Frame(root)
main_frame.pack(fill="both", expand=True)

# Left menu frame
menu_frame = tk.Frame(main_frame, width=150, bg="#3C3F41", relief="raised")
menu_frame.pack(side="left", fill="y")

# Right content frame
content_frame = tk.Frame(main_frame, bg="#F5F5F5")
content_frame.pack(side="right", expand=True, fill="both")

# Bottom bar
bottom_bar = tk.Frame(root, bg="#2E2E2E", height=50)
bottom_bar.pack(side="bottom", fill="x")

# Footer label with dynamic text
footer_label = tk.Label(
    bottom_bar,
    text=f"© Voyago | {datetime.now().strftime('%Y-%m-%d')}",
    bg="#2E2E2E",
    fg="white",
    font=("Arial", 10),
)
footer_label.pack(side="right", padx=10)

# Define frames for each screen in the content area
clean_frame = tk.Frame(content_frame, bg="#F5F5F5")
transform_frame = tk.Frame(content_frame, bg="#F5F5F5")
transfer_frame = tk.Frame(content_frame, bg="#F5F5F5")
compare_frame = tk.Frame(content_frame, bg="#F5F5F5")

# Place all frames on the same stack
for frame in (clean_frame, transform_frame, transfer_frame, compare_frame):
    frame.place(relwidth=1, relheight=1)

# region Clean Screen widgets
# Add widgets to the Clean Screen
tk.Label(clean_frame, text="Clean Report", bg="#F5F5F5", font=("Arial", 16)).pack(pady=10)

label_select_report = tk.Label(clean_frame, text="Select Report to Clean", font=("Arial", 12), bg="#F5F5F5")
label_select_report.pack(pady=5)

selected_report = tk.StringVar(value="Deficiency_Recertification")  # Default report selection

radio_deficiency = tk.Radiobutton(
    clean_frame, text="ADP Deficiency_Recertification Report", variable=selected_report,
    value="Deficiency_Recertification", bg="#F5F5F5", font=("Arial", 10)
)
radio_deficiency.pack(anchor="w", padx=(50, 0), pady=(10, 0))

radio_policies = tk.Radiobutton(
    clean_frame, text="ADP Policies_Certifications_Vaccines_Licences Report", variable=selected_report,
    value="Policies_Certifications_Vaccines_Licences", bg="#F5F5F5", font=("Arial", 10)
)
radio_policies.pack(anchor="w", padx=(50, 0), pady=(0, 0))

radio_courses = tk.Radiobutton(
    clean_frame, text="ADP All_Course_Progresses Report", variable=selected_report,
    value="All_Course_Progresses", bg="#F5F5F5", font=("Arial", 10)
)
radio_courses.pack(anchor="w", padx=(50, 0), pady=(0, 10))

clean_browse_button = tk.Button(clean_frame, text="Select Report", font=("Arial", 12),
                                width=25, height=1, command=select_clean_file)
clean_browse_button.pack(pady=5)

file_label = tk.Label(clean_frame, text="No file selected", bg="#F5F5F5", font=("Arial", 10), wraplength=400)
file_label.pack(pady=5)

start_button = tk.Button(clean_frame, text="Start Clean", font=("Arial", 14),
                         width=20, height=2, command=start_clean_logic)
start_button.pack(pady=10)
# endregion

# region Transform Screen widgets
# Add widgets to the Transform Screen
tk.Label(transform_frame, text="Transform Report", bg="#F5F5F5", font=("Arial", 16)).pack(pady=10)

transform_browse_button = tk.Button(transform_frame, text="Select Cleaned Report", font=("Arial", 12),
                                           width=25, height=1, command=select_transform_file)
transform_browse_button.pack(pady=5)

transform_file_label = tk.Label(transform_frame, text="No file selected", bg="#F5F5F5", font=("Arial", 10), wraplength=400)
transform_file_label.pack(pady=5)

transform_course_mapping_button = tk.Button(transform_frame, text="Add Course Mapping", font=("Arial", 12),
                                            width=25, height=1, command=select_course_mapping_file)
transform_course_mapping_button.pack(pady=5)

course_mapping_file_label = tk.Label(transform_frame, text="No file selected", bg="#F5F5F5", font=("Arial", 10), wraplength=400)
course_mapping_file_label.pack(pady=5)

transform_user_list_button = tk.Button(transform_frame, text="Add User List", font=("Arial", 12),
                                       width=25, height=1, command=select_user_list_file)
transform_user_list_button.pack(pady=5)

user_list_file_label = tk.Label(transform_frame, text="No file selected", bg="#F5F5F5", font=("Arial", 10), wraplength=400)
user_list_file_label.pack(pady=5)

start_transform_button = tk.Button(transform_frame, text="Start Transform", font=("Arial", 14),
                                   width=20, height=2, command=start_transform_logic)
start_transform_button.pack(pady=10)
# endregion

# region Transfer Screen widgets
# Add widgets to the Transfer Screen
tk.Label(transfer_frame, text="Transfer Report", bg="#F5F5F5", font=("Arial", 16)).pack(pady=30)

transfer_browse_button = tk.Button(transfer_frame, text="Select Output Report", font=("Arial", 12),
                                   width=25, height=1, command=select_transfer_file)
transfer_browse_button.pack(pady=5)

transfer_file_label = tk.Label(transfer_frame, text="No file selected", bg="#F5F5F5", font=("Arial", 10), wraplength=400)
transfer_file_label.pack(pady=5)

start_transfer_button = tk.Button(transfer_frame, text="Start Transfer", font=("Arial", 14),
                                  width=20, height=2, command=start_transfer_logic)
start_transfer_button.pack(pady=30)
# endregion

# region Compare Screen widgets
# Add widgets to the Compare Screen
tk.Label(compare_frame, text="Compare Reports", bg="#F5F5F5", font=("Arial", 16)).pack(pady=20)

compare_browse_button = tk.Button(compare_frame, text="Select Generated Report", font=("Arial", 12),
                                  width=25, height=1, command=select_compare_file)
compare_browse_button.pack(pady=5)

compare_file_label = tk.Label(compare_frame, text="No file selected", bg="#F5F5F5", font=("Arial", 10), wraplength=400)
compare_file_label.pack(pady=5)

reference_browse_button = tk.Button(compare_frame, text="Select Reference Report", font=("Arial", 12),
                                    width=25, height=1, command=select_reference_file)
reference_browse_button.pack(pady=5)

reference_file_label = tk.Label(compare_frame, text="No file selected", bg="#F5F5F5", font=("Arial", 10), wraplength=400)
reference_file_label.pack(pady=5)

start_compare_button = tk.Button(compare_frame, text="Start Compare", font=("Arial", 14),
                                 width=20, height=2, command=start_compare_logic)
start_compare_button.pack(pady=30)
# endregion

# Bring the selected frame to the front
def show_frame(frame):
    frame.tkraise()

# Change button color on hover to a darker version
def on_enter(e):    
    idx = button_widgets.index(e.widget)
    original_color = buttons[idx][1]
    # Darken the color slightly
    darker_color = "#%02x%02x%02x" % tuple(max(0, int(original_color[i:i+2], 16) - 30) for i in (1, 3, 5))
    e.widget['bg'] = darker_color

# Revert button color when hover ends
def on_leave(e):    
    idx = button_widgets.index(e.widget)
    e.widget['bg'] = buttons[idx][1]  # Original color

# Function to dynamically resize buttons with spacing and padding
def resize_buttons():    
    frame_width = menu_frame.winfo_width()
    button_width = frame_width - 2 * padding  # Button width matches the menu frame width with padding
    button_height = 60  # Fixed height for all buttons

    for idx, button in enumerate(button_widgets):
        button.place(
            x=padding,  # Align with padding
            y=padding + idx * (button_height + spacing),  # Space buttons vertically
            width=button_width,
            height=button_height,
        )

# Define button properties
buttons = [(text, "#E90000", frame) for text, frame in [
    ("Clean", clean_frame),
    ("Transform", transform_frame),
    ("Transfer", transfer_frame),
    ("Compare", compare_frame),
]]
button_widgets = []
padding = 20  # Padding around the buttons
spacing = 30  # Space between buttons

# Add buttons to the menu frame with hover effects
for text, color, frame in buttons:
    btn = tk.Button(
        menu_frame,
        text=text,
        bg=color,
        fg="white",
        font=("Arial", 12, "bold"),
        relief="raised",
        borderwidth=2,
        command=lambda f=frame: show_frame(f),
    )
    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    button_widgets.append(btn)

# Bind the resize event to dynamically adjust button size, padding and spacing
menu_frame.bind("<Configure>", lambda e: resize_buttons())

# Show the first screen by default
show_frame(clean_frame)

# Run the application
root.mainloop()

# endregion